from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple
import re

import openpyxl
import yaml

from parser.common import normalize_scalar, write_json
from parser.excel_parser import normalize_header

PARSER_VERSION = "2.0.0-m7.1"


def _safe_filename(value: str) -> str:
    value = re.sub(r"[\\/:*?\"<>|\x00-\x1f]", "_", value.strip())
    return value[:160] or "UNKNOWN"


@dataclass
class QueryParseSummary:
    source_excel: str
    sheet_name: str = ""
    header_row: int = 0
    header_rows: int = 1
    headers: List[str] = field(default_factory=list)
    sheet_diagnostics: List[dict] = field(default_factory=list)
    missing_columns: List[str] = field(default_factory=list)
    total: int = 0
    success: int = 0
    partial_success: int = 0
    failed: int = 0
    duplicate_query_ids: List[str] = field(default_factory=list)
    failed_rows: List[dict] = field(default_factory=list)
    generated_at: str = ""

    def to_dict(self) -> dict:
        return self.__dict__.copy()


class QueryExcelParser:
    def __init__(self, mapping_config: str | Path) -> None:
        config = yaml.safe_load(Path(mapping_config).read_text(encoding="utf-8"))["query_excel"]
        self.sheet_name = config.get("sheet_name")
        self.header_row = int(config["header_row"]) if config.get("header_row") not in (None, "") else None
        self.header_rows_mode = config.get("header_rows", "auto")
        self.fields: Dict[str, str] = config["fields"]
        self.required_fields = list(config.get("required_fields", ["query_id", "problem_description"]))
        auto = config.get("auto_detect", {})
        self.scan_rows = int(auto.get("scan_rows", 20))
        self.scan_columns = int(auto.get("scan_columns", 64))
        self.minimum_required_matches = int(auto.get("minimum_required_matches", 2))
        self.preferred_fields = [normalize_header(x) for x in auto.get("preferred_fields", [])]
        self.alias_lookup: Dict[str, str] = {}
        for canonical, aliases in config.get("column_aliases", {}).items():
            canonical_n = normalize_header(canonical)
            for alias in [canonical, *(aliases or [])]:
                self.alias_lookup[normalize_header(alias)] = canonical_n
        for value in self.fields.values():
            n = normalize_header(value)
            self.alias_lookup.setdefault(n, n)

    def _canonical(self, value: Any) -> str:
        n = normalize_header(value)
        return self.alias_lookup.get(n, n)

    def _raw_row(self, ws, row: int) -> List[str]:
        cells = next(ws.iter_rows(min_row=row, max_row=row, min_col=1, max_col=max(self.scan_columns, int(ws.max_column or 1))), ())
        values = [normalize_scalar(c.value) for c in cells]
        while values and values[-1] == "": values.pop()
        return values

    def _headers(self, ws, row: int, count: int) -> Tuple[List[str], List[str]]:
        first = self._raw_row(ws, row)
        if count == 1:
            display = first
        else:
            second = self._raw_row(ws, row + 1)
            width = max(len(first), len(second))
            display = []
            for i in range(width):
                a = first[i] if i < len(first) else ""
                b = second[i] if i < len(second) else ""
                display.append(b or a)
        canonical = [self._canonical(x) for x in display]
        return display, canonical

    def _score(self, headers: List[str]) -> Tuple[int, int]:
        values = {x for x in headers if x}
        mapped = {self._canonical(v) for v in self.fields.values()}
        return sum(x in values for x in self.preferred_fields), sum(x in values for x in mapped)

    def _detect(self, wb):
        sheets = [wb[self.sheet_name]] if self.sheet_name else list(wb.worksheets)
        candidates = []
        diagnostics = []
        header_counts = [int(self.header_rows_mode)] if str(self.header_rows_mode) in {"1", "2"} else [1, 2]
        for order, ws in enumerate(sheets):
            best = None
            max_row = min(max(int(ws.max_row or 1), 1), self.scan_rows)
            rows = [self.header_row] if self.header_row else range(1, max_row + 1)
            for row in rows:
                for count in header_counts:
                    if row + count - 1 > int(ws.max_row or 1): continue
                    display, canonical = self._headers(ws, row, count)
                    preferred, total = self._score(canonical)
                    candidate = (preferred, total, -row, -count, -order, ws, row, count, display, canonical)
                    candidates.append(candidate)
                    if best is None or candidate[:5] > best[:5]: best = candidate
            if best:
                diagnostics.append({"sheet_name": ws.title, "max_row": ws.max_row, "max_column": ws.max_column,
                    "best_header_row": best[6], "header_rows": best[7], "preferred_field_matches": best[0],
                    "mapped_field_matches": best[1], "headers": best[8]})
        if not candidates: raise ValueError("Excel中没有可读取的工作表")
        best = max(candidates, key=lambda x: x[:5])
        if max(best[0], best[1]) < self.minimum_required_matches:
            raise ValueError(f"未识别到有效Query表头，至少需匹配{self.minimum_required_matches}个字段: {diagnostics}")
        return best[5], best[6], best[7], best[8], best[9], diagnostics

    def parse(self, excel_path: str | Path, output_dir: str | Path) -> Tuple[List[dict], QueryParseSummary]:
        source = Path(excel_path)
        if not source.exists(): raise FileNotFoundError(f"Excel文件不存在: {source}")
        wb = openpyxl.load_workbook(source, data_only=True, read_only=False)
        try:
            ws, header_row, header_rows, display_headers, headers, diagnostics = self._detect(wb)
            index = {h: i for i, h in enumerate(headers) if h and h not in headers[:i]}
            canonical_mapping = {k: self._canonical(v) for k, v in self.fields.items()}
            missing_columns = sorted(self.fields[k] for k, v in canonical_mapping.items() if v not in index)
            summary = QueryParseSummary(str(source), ws.title, header_row, header_rows, display_headers, diagnostics, missing_columns)
            out = Path(output_dir); out.mkdir(parents=True, exist_ok=True)
            records, seen = [], set()
            max_col = max(len(headers), int(ws.max_column or 1), self.scan_columns)
            start = header_row + header_rows
            for excel_row, row in enumerate(ws.iter_rows(min_row=start, min_col=1, max_col=max_col), start=start):
                values = [normalize_scalar(c.value) for c in row]
                if not any(v != "" for v in values): continue
                raw_fields = {display_headers[i] or f"COLUMN_{i+1}": values[i] if i < len(values) else "" for i in range(len(display_headers))}
                mapped = {k: (values[index[v]] if v in index and index[v] < len(values) else "") for k, v in canonical_mapping.items()}
                query_id = str(mapped.pop("query_id", "")).strip()
                warnings = [f"MISSING_COLUMN:{x}" for x in missing_columns]
                missing_required = []
                for field_name in self.required_fields:
                    val = query_id if field_name == "query_id" else str(mapped.get(field_name, "")).strip()
                    if not val: missing_required.append(field_name)
                if query_id and query_id in seen:
                    warnings.append(f"DUPLICATE_QUERY_ID:{query_id}"); missing_required.append("query_id_duplicate")
                    summary.duplicate_query_ids.append(query_id)
                if query_id: seen.add(query_id)
                if missing_required:
                    warnings.extend(f"MISSING_REQUIRED:{x}" for x in missing_required)
                    status = "QUERY_PARSE_FAILED"
                elif warnings:
                    status = "PARTIAL_SUCCESS"
                else:
                    status = "SUCCESS"
                record = {"query_id": query_id, "source_excel": str(source), "sheet_name": ws.title, "excel_row": excel_row,
                    "raw_fields": raw_fields, "mapped_fields": mapped, "parse_status": status,
                    "parse_warnings": sorted(set(warnings)), "parser_version": PARSER_VERSION,
                    "generated_at": datetime.now(timezone.utc).isoformat()}
                filename = _safe_filename(query_id) if query_id and status != "QUERY_PARSE_FAILED" else f"FAILED-ROW-{excel_row}"
                write_json(out / f"{filename}.json", record)
                records.append(record); summary.total += 1
                if status == "SUCCESS": summary.success += 1
                elif status == "PARTIAL_SUCCESS": summary.partial_success += 1
                else:
                    summary.failed += 1
                    summary.failed_rows.append({"excel_row": excel_row, "query_id": query_id, "warnings": record["parse_warnings"]})
            if summary.total == 0: raise ValueError("已识别Query表头，但未读取到数据行")
            summary.duplicate_query_ids = sorted(set(summary.duplicate_query_ids))
            summary.generated_at = datetime.now(timezone.utc).isoformat()
            return records, summary
        finally:
            wb.close()
