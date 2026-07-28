from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple
import re

import openpyxl
import yaml

from parser.common import normalize_scalar, write_json


def normalize_header(value: Any) -> str:
    """Normalize visually identical Excel headers for stable matching."""
    text = normalize_scalar(value)
    text = text.replace("\u3000", " ").replace("\xa0", " ")
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[\r\n\t ]+", "", text)
    return text.strip()


@dataclass
class ExcelParseSummary:
    source_excel: str
    sheet_name: str
    header_row: int
    headers: List[str]
    sheet_diagnostics: List[dict]
    missing_columns: List[str]
    total_rows: int = 0
    success_count: int = 0
    partial_success_count: int = 0
    failed_count: int = 0
    duplicate_itr_ids: List[str] = None

    def __post_init__(self) -> None:
        if self.duplicate_itr_ids is None:
            self.duplicate_itr_ids = []

    def to_dict(self) -> dict:
        return {
            "source_excel": self.source_excel,
            "sheet_name": self.sheet_name,
            "header_row": self.header_row,
            "headers": self.headers,
            "sheet_diagnostics": self.sheet_diagnostics,
            "missing_columns": self.missing_columns,
            "total_rows": self.total_rows,
            "success_count": self.success_count,
            "partial_success_count": self.partial_success_count,
            "failed_count": self.failed_count,
            "duplicate_itr_ids": self.duplicate_itr_ids,
        }


class ExcelParser:
    def __init__(
        self,
        mapping_config: str | Path,
        case_id_prefix: str = "CASE",
        case_id_width: int = 6,
    ) -> None:
        config = yaml.safe_load(Path(mapping_config).read_text(encoding="utf-8"))
        excel_cfg = config["excel"]
        self.sheet_name = excel_cfg.get("sheet_name")
        configured_header = excel_cfg.get("header_row")
        self.header_row = int(configured_header) if configured_header not in (None, "") else None
        self.field_mapping: Dict[str, str] = excel_cfg["fields"]
        self.case_id_prefix = case_id_prefix
        self.case_id_width = case_id_width

        auto_cfg = excel_cfg.get("auto_detect", {})
        self.auto_detect_enabled = bool(auto_cfg.get("enabled", True))
        self.scan_rows = int(auto_cfg.get("scan_rows", 20))
        self.minimum_required_matches = int(auto_cfg.get("minimum_required_matches", 2))
        self.preferred_fields = [
            normalize_header(item)
            for item in auto_cfg.get(
                "preferred_fields",
                ["ITR单号", "问题描述", "IPMT", "SPDT", "复盘报告"],
            )
        ]

        aliases = excel_cfg.get("column_aliases", {})
        self.alias_lookup: Dict[str, str] = {}
        for canonical, alias_values in aliases.items():
            canonical_normalized = normalize_header(canonical)
            for alias in [canonical, *(alias_values or [])]:
                self.alias_lookup[normalize_header(alias)] = canonical_normalized

        for column_name in self.field_mapping.values():
            normalized = normalize_header(column_name)
            self.alias_lookup.setdefault(normalized, normalized)

    def _canonical_header(self, value: Any) -> str:
        normalized = normalize_header(value)
        return self.alias_lookup.get(normalized, normalized)

    def _read_row_headers(self, worksheet, row_number: int) -> List[str]:
        # Some exported workbooks contain an incorrect dimension such as A1:A23,
        # although visible data exists in many columns. Do not trust max_column.
        configured_columns = max(len(self.field_mapping) + 10, 64)
        detected_columns = max(int(getattr(worksheet, "max_column", 1) or 1), configured_columns)
        row = next(
            worksheet.iter_rows(
                min_row=row_number,
                max_row=row_number,
                min_col=1,
                max_col=detected_columns,
            ),
            (),
        )
        headers = [self._canonical_header(cell.value) for cell in row]
        while headers and not headers[-1]:
            headers.pop()
        return headers

    def _score_header(self, headers: List[str]) -> tuple[int, int]:
        nonempty = {header for header in headers if header}
        mapped_columns = {
            self._canonical_header(value)
            for value in self.field_mapping.values()
        }
        preferred_matches = sum(1 for field in self.preferred_fields if field in nonempty)
        total_matches = sum(1 for field in mapped_columns if field in nonempty)
        return preferred_matches, total_matches

    def _detect_sheet_and_header(self, workbook) -> tuple[Any, int, List[str], List[dict]]:
        diagnostics: List[dict] = []
        candidates: List[tuple[int, int, int, Any, int, List[str]]] = []

        worksheets = (
            [workbook[self.sheet_name]]
            if self.sheet_name
            else list(workbook.worksheets)
        )

        for sheet_order, worksheet in enumerate(worksheets):
            max_scan = min(max(worksheet.max_row or 1, 1), self.scan_rows)
            best_sheet_candidate = None

            rows_to_scan = (
                [self.header_row]
                if self.header_row is not None
                else range(1, max_scan + 1)
            )
            for row_number in rows_to_scan:
                headers = self._read_row_headers(worksheet, row_number)
                preferred_matches, total_matches = self._score_header(headers)
                candidate = (
                    preferred_matches,
                    total_matches,
                    -sheet_order,
                    worksheet,
                    row_number,
                    headers,
                )
                if best_sheet_candidate is None or candidate[:3] > best_sheet_candidate[:3]:
                    best_sheet_candidate = candidate
                candidates.append(candidate)

            if best_sheet_candidate is not None:
                diagnostics.append({
                    "sheet_name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "best_header_row": best_sheet_candidate[4],
                    "preferred_field_matches": best_sheet_candidate[0],
                    "mapped_field_matches": best_sheet_candidate[1],
                    "headers": best_sheet_candidate[5],
                })

        if not candidates:
            raise ValueError("Excel中没有可读取的工作表")

        best = max(candidates, key=lambda item: item[:3])
        preferred_matches, total_matches, _, worksheet, header_row, headers = best

        if self.auto_detect_enabled:
            if (
                preferred_matches < self.minimum_required_matches
                and total_matches < self.minimum_required_matches
            ):
                detail = "; ".join(
                    f"{item['sheet_name']}[row={item['best_header_row']},"
                    f"matched={item['mapped_field_matches']},headers={item['headers']}]"
                    for item in diagnostics
                )
                raise ValueError(
                    "未识别到有效Excel表头。"
                    f"至少需要匹配{self.minimum_required_matches}个字段。扫描结果: {detail}"
                )

        return worksheet, header_row, headers, diagnostics

    def parse(
        self,
        excel_path: str | Path,
        output_dir: str | Path,
    ) -> Tuple[List[dict], ExcelParseSummary]:
        source = Path(excel_path)
        if not source.exists():
            raise FileNotFoundError(f"Excel文件不存在: {source}")

        workbook = openpyxl.load_workbook(source, data_only=True, read_only=False)
        try:
            worksheet, header_row, headers, diagnostics = self._detect_sheet_and_header(workbook)
            header_index = {
                name: index
                for index, name in enumerate(headers)
                if name and name not in headers[:index]
            }

            canonical_mapping = {
                standard_field: self._canonical_header(column_name)
                for standard_field, column_name in self.field_mapping.items()
            }
            missing_columns = sorted({
                original_column
                for standard_field, original_column in self.field_mapping.items()
                if canonical_mapping[standard_field] not in header_index
            })

            records: List[dict] = []
            seen_itr: Dict[str, int] = {}
            duplicate_itr_ids: set[str] = set()
            summary = ExcelParseSummary(
                source_excel=str(source),
                sheet_name=worksheet.title,
                header_row=header_row,
                headers=headers,
                sheet_diagnostics=diagnostics,
                missing_columns=missing_columns,
            )

            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)

            data_sequence = 0
            data_max_column = max(
                len(headers),
                int(getattr(worksheet, "max_column", 1) or 1),
                64,
            )
            for excel_row, row in enumerate(
                worksheet.iter_rows(
                    min_row=header_row + 1,
                    min_col=1,
                    max_col=data_max_column,
                ),
                start=header_row + 1,
            ):
                raw_values = [normalize_scalar(cell.value) for cell in row]
                if not any(value != "" for value in raw_values):
                    continue

                data_sequence += 1
                case_id = f"{self.case_id_prefix}-{data_sequence:0{self.case_id_width}d}"
                raw_fields = {
                    header: raw_values[index] if index < len(raw_values) else ""
                    for index, header in enumerate(headers)
                    if header
                }
                mapped_fields: Dict[str, Any] = {}
                warnings: List[str] = []

                for standard_field, canonical_column in canonical_mapping.items():
                    if canonical_column not in header_index:
                        mapped_fields[standard_field] = ""
                        warnings.append(
                            f"MISSING_COLUMN:{self.field_mapping[standard_field]}"
                        )
                        continue
                    index = header_index[canonical_column]
                    mapped_fields[standard_field] = (
                        raw_values[index] if index < len(raw_values) else ""
                    )

                itr_id = str(mapped_fields.get("itr_id", "")).strip()
                if not itr_id:
                    warnings.append("MISSING_ITR_ID")
                elif itr_id in seen_itr:
                    duplicate_itr_ids.add(itr_id)
                    warnings.append(f"DUPLICATE_ITR_ID:{itr_id}")
                else:
                    seen_itr[itr_id] = excel_row

                status = "SUCCESS" if not warnings else "PARTIAL_SUCCESS"
                record = {
                    "case_id": case_id,
                    "source_excel": str(source),
                    "sheet_name": worksheet.title,
                    "excel_row": excel_row,
                    "raw_fields": raw_fields,
                    "mapped_fields": mapped_fields,
                    "parse_status": status,
                    "parse_warnings": sorted(set(warnings)),
                    "generated_at": datetime.now(timezone.utc).isoformat(),
                }
                write_json(output_path / f"{case_id}.json", record)
                records.append(record)

                summary.total_rows += 1
                if status == "SUCCESS":
                    summary.success_count += 1
                elif status == "PARTIAL_SUCCESS":
                    summary.partial_success_count += 1
                else:
                    summary.failed_count += 1

            summary.duplicate_itr_ids = sorted(duplicate_itr_ids)

            if summary.total_rows == 0:
                raise ValueError(
                    "已识别Sheet和表头，但未读取到任何数据行。"
                    f"sheet={worksheet.title}, header_row={header_row}, "
                    f"max_row={worksheet.max_row}, headers={headers}"
                )

            return records, summary
        finally:
            workbook.close()
